"""Normaliza as fontes oficiais da pasta KB em datasets JSON usados pelo app.

Uso:
    python scripts/build_dataset.py

Entradas (KB/):
    nfse-rtc/AnexoVIII-...xlsx      Correlacao Item LC116 x NBS x indOP x cClassTrib
    nfse-rtc/AnexoVII-IndOp...xlsx  Tabela de indicadores de operacao (cIndOp)
    nfe-rtc/Tabela_Classificacao_Tributaria_IBS_CBS_*.xlsx
    nfe-rtc/Tabela_Codigos_Item_Lista_Servicos.xlsx
    nbs/NBS_2.0_tabela.csv
    cnae/ibge_cnae_subclasses.json
    cnae-lc116/*.pdf                Tabelas municipais CNAE x item LC116

Saidas (data/):
    lc116.json, nbs.json, indop.json, cclasstrib.json, cst.json,
    anexo_viii.json, cnae.json, cnae_lc116.json, manifesto.json
"""

from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import re
import sys
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[1]
KB = RAIZ / "KB"
DATA = RAIZ / "data"


def _txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(r"\s+", " ", str(v)).strip()


def _sha256(caminho: Path) -> str:
    h = hashlib.sha256()
    with caminho.open("rb") as fh:
        for bloco in iter(lambda: fh.read(1 << 20), b""):
            h.update(bloco)
    return h.hexdigest()


def _unico(caminho_glob: str) -> Path:
    achados = sorted(KB.glob(caminho_glob))
    if not achados:
        raise SystemExit(f"Fonte nao encontrada em KB: {caminho_glob}")
    return achados[-1]


# ---------------------------------------------------------------- LC 116/2003


def build_lc116() -> list[dict]:
    arq = _unico("nfe-rtc/Tabela_Codigos_Item_Lista_Servicos.xlsx")
    ws = openpyxl.load_workbook(arq, read_only=True).worksheets[0]
    itens = []
    for linha in ws.iter_rows(values_only=True):
        codigo, desc = _txt(linha[0]), _txt(linha[1] if len(linha) > 1 else "")
        if not re.fullmatch(r"\d{2}\.\d{2}", codigo):
            continue
        itens.append({"item": codigo, "descricao": desc})
    return itens


# ----------------------------------------------------------------------- NBS


def build_nbs() -> list[dict]:
    arq = _unico("nbs/NBS_2.0_tabela.csv")
    texto = arq.read_bytes().decode("latin-1")
    itens = []
    for linha in csv.reader(io.StringIO(texto), delimiter=";"):
        if len(linha) < 2:
            continue
        codigo, desc = _txt(linha[0]), _txt(linha[1])
        if not re.fullmatch(r"[\d.]+", codigo):
            continue
        itens.append(
            {
                "nbs": codigo,
                "descricao": desc,
                "nivel": "item" if codigo.count(".") == 3 else "agrupamento",
            }
        )
    return itens


# ---------------------------------------------------------------------- NEBS


# cabecalho de nota: "1.0102.61 Servicos de construcao de usinas..."
RE_NEBS = re.compile(r"^(\d\.\d{2}(?:\d{2})?(?:\.\d{1,2}){0,2})\s+(\S.*)$")
# inicios tipicos do corpo da nota (nao sao continuacao de titulo)
RE_CORPO = re.compile(
    r"^(Est[aeã]\w*|Inclui|Exclu\w+|Compreende|Observa\w+|[-–•]|\d+\s*[-–])",
    re.IGNORECASE,
)


def build_nebs() -> list[dict]:
    """Le as Notas Explicativas (NEBS) do PDF, uma entrada por codigo NBS.

    O texto de cada codigo vai do seu cabecalho ate o cabecalho seguinte. Os
    codigos do PDF vao ate a subposicao (ex.: 1.0102.61), enquanto a tabela da
    NBS usa o item completo (1.0102.61.00) - a ligacao entre os dois e feita na
    consulta, por prefixo.
    """
    import pdfplumber

    arq = _unico("nbs/NEBS_2.0_AnexoII_notas_explicativas.pdf")
    notas: list[dict] = []
    atual: dict | None = None

    with pdfplumber.open(arq) as pdf:
        for pagina in pdf.pages:
            for linha in (pagina.extract_text() or "").splitlines():
                linha = linha.strip()
                if not linha or linha.startswith("(Fl."):
                    continue

                achado = RE_NEBS.match(linha)
                # um codigo citado no meio da frase pode cair no inicio da linha
                # seguinte; cabecalho de verdade sempre abre com maiuscula
                if achado and not achado.group(2)[:1].isupper():
                    achado = None
                if achado:
                    atual = {
                        "nbs": achado.group(1),
                        "titulo": _txt(achado.group(2)),
                        "paragrafos": [],
                        "pagina": pagina.page_number,
                    }
                    notas.append(atual)
                    continue

                if atual is None:
                    continue  # texto de apresentacao, antes da primeira nota

                # titulo quebrado em duas linhas: completa o cabecalho
                if not atual["paragrafos"] and not RE_CORPO.match(linha) and len(linha) < 70:
                    atual["titulo"] = _txt(f"{atual['titulo']} {linha}")
                    continue

                atual["paragrafos"].append(linha)

    for nota in notas:
        nota["texto"] = "\n".join(nota.pop("paragrafos")).strip()

    return [n for n in notas if n["texto"]]


# --------------------------------------------------------------- indOP / RTC


def build_indop() -> list[dict]:
    arq = _unico("nfse-rtc/AnexoVII-IndOp_IBSCBS_*.xlsx")
    ws = openpyxl.load_workbook(arq, read_only=True).worksheets[0]
    linhas = list(ws.iter_rows(values_only=True))
    itens = []
    for linha in linhas[1:]:
        codigo = _txt(linha[0])
        if not re.fullmatch(r"\d{6}", codigo):
            continue
        itens.append(
            {
                "indop": codigo,
                "tipo_operacao": _txt(linha[1]),
                "caracteristica": _txt(linha[2]),
                "local_fornecimento": _txt(linha[3]),
                "dispositivo_lc214": _txt(linha[4]),
                "observacao": _txt(linha[5]),
                "usa_nfe": _txt(linha[6]).upper() == "S",
                "usa_nfse": _txt(linha[7]).upper() == "S",
            }
        )
    return itens


def build_cclasstrib() -> tuple[list[dict], list[dict]]:
    arq = _unico("nfe-rtc/Tabela_Classificacao_Tributaria_IBS_CBS_*.xlsx")
    wb = openpyxl.load_workbook(arq, read_only=True)

    aba_cst = next(ws for ws in wb.worksheets if ws.title.lower().startswith("cst"))
    cst = []
    for linha in list(aba_cst.iter_rows(values_only=True))[1:]:
        codigo = _txt(linha[0])
        if not re.fullmatch(r"\d{3}", codigo):
            continue
        cst.append({"cst": codigo, "descricao": _txt(linha[1])})

    aba = next(ws for ws in wb.worksheets if ws.title.lower().startswith("cclass"))
    linhas = list(aba.iter_rows(values_only=True))
    cabecalho = [_txt(c) for c in linhas[0]]
    idx = {nome: i for i, nome in enumerate(cabecalho) if nome}

    def col(linha, nome, padrao=""):
        i = idx.get(nome)
        return _txt(linha[i]) if i is not None and i < len(linha) else padrao

    classes = []
    for linha in linhas[1:]:
        codigo = col(linha, "cClassTrib")
        if not re.fullmatch(r"\d{6}", codigo):
            continue
        classes.append(
            {
                "cclasstrib": codigo,
                "cst": col(linha, "CST-IBS/CBS"),
                "cst_descricao": col(linha, "Descrição CST-IBS/CBS"),
                "nome": col(linha, "Nome cClassTrib"),
                "descricao": col(linha, "Descrição cClassTrib"),
                "lc214": col(linha, "LC 214/25"),
                "tipo_aliquota": col(linha, "Tipo de Alíquota"),
                "p_red_ibs": col(linha, "pRedIBS"),
                "p_red_cbs": col(linha, "pRedCBS"),
                "vigencia_inicio": col(linha, "dIniVig")[:10],
                "vigencia_fim": col(linha, "dFimVig")[:10],
                "atualizado_em": col(linha, "DataAtualização")[:10],
                "usa_nfe": col(linha, "indNFe") == "1",
                "usa_nfce": col(linha, "indNFCe") == "1",
                "usa_nfse": col(linha, "indNFSe") == "1",
                "usa_nfcom": col(linha, "indNFCom") == "1",
                "usa_cte": col(linha, "indCTe") == "1",
                "anexo_lc214": col(linha, "ANEXO"),
                "link": col(linha, "Link"),
            }
        )
    return classes, cst


# ----------------------------------------------------------------- Anexo VIII


def _linhas_com_mescla(arq: Path, filtro_aba) -> list[list[str]]:
    """Le uma aba propagando o valor das celulas mescladas para todas as linhas."""
    wb = openpyxl.load_workbook(arq)
    aba = next(ws for ws in wb.worksheets if filtro_aba(ws.title))
    grade = [[_txt(c.value) for c in linha] for linha in aba.iter_rows()]
    for faixa in aba.merged_cells.ranges:
        valor = _txt(aba.cell(faixa.min_row, faixa.min_col).value)
        if not valor:
            continue
        for r in range(faixa.min_row, faixa.max_row + 1):
            for c in range(faixa.min_col, faixa.max_col + 1):
                grade[r - 1][c - 1] = valor
    return grade


def build_anexo_viii() -> list[dict]:
    """Le a aba 'tabela geral' do Anexo VIII.

    A planilha usa celulas mescladas: item da LC 116, NBS, cClassTrib, local de
    incidencia e onerosidade valem para todas as linhas da mescla. Propagados esses
    valores, cada linha vira a tupla completa
    (item, NBS, indOP, local, onerosidade, cClassTrib).

    O resultado agrupa, por item e cClassTrib, as NBS que compartilham o mesmo
    conjunto de indOP - assim um item de saude, por exemplo, aparece uma vez com os
    tres indicadores aplicaveis (presencial, nao presencial e a distancia).
    """
    arq = _unico("nfse-rtc/AnexoVIII-CorrelacaoItemNBSIndOpCClassTrib_IBSCBS_*.xlsx")
    grade = _linhas_com_mescla(arq, lambda t: "geral" in t.lower())

    # (item, cClassTrib, NBS) -> dados agregados
    combinacoes: dict[tuple, dict] = {}
    for linha in grade[1:]:
        c = (linha + [""] * 10)[:10]
        item, item_desc, nbs, nbs_desc = c[0], c[1], c[2], c[3]
        ps, adq, indop, local, cclass, cclass_nome = c[4:10]
        if not item or not cclass:
            continue
        chave = (item, cclass, nbs)
        registro = combinacoes.setdefault(
            chave,
            {
                "item": item,
                "item_descricao": item_desc,
                "cclasstrib": cclass,
                "cclasstrib_nome": cclass_nome,
                "nbs": nbs,
                "nbs_descricao": nbs_desc,
                "indops": [],
            },
        )
        if indop and indop not in [i["indop"] for i in registro["indops"]]:
            registro["indops"].append(
                {
                    "indop": indop,
                    "local_incidencia": local,
                    "ps_onerosa": ps,
                    "adquirido_exterior": adq,
                }
            )

    # agrupa NBS que compartilham item, cClassTrib e conjunto de indOP
    agrupado: dict[tuple, dict] = {}
    for registro in combinacoes.values():
        chave = (
            registro["item"],
            registro["cclasstrib"],
            tuple(sorted(i["indop"] for i in registro["indops"])),
        )
        alvo = agrupado.get(chave)
        if alvo is None:
            alvo = agrupado[chave] = {
                "item": registro["item"],
                "item_descricao": registro["item_descricao"],
                "cclasstrib": registro["cclasstrib"],
                "cclasstrib_nome": registro["cclasstrib_nome"],
                "indops": registro["indops"],
                "local_incidencia": " / ".join(
                    dict.fromkeys(
                        i["local_incidencia"] for i in registro["indops"] if i["local_incidencia"]
                    )
                ),
                "nbs": [],
            }
        if registro["nbs"] and registro["nbs"] not in [n["nbs"] for n in alvo["nbs"]]:
            alvo["nbs"].append(
                {"nbs": registro["nbs"], "descricao": registro["nbs_descricao"]}
            )

    return list(agrupado.values())


# ----------------------------------------------------------------------- NCM


def build_ncm() -> list[dict]:
    """Nomenclatura Comum do Mercosul (tabela do Portal Unico Siscomex)."""
    arq = _unico("ncm/siscomex_nomenclatura_ncm.json")
    bruto = json.loads(arq.read_text(encoding="utf-8"))
    hoje = datetime.now().date()
    itens = []
    for registro in bruto.get("Nomenclaturas", []):
        codigo = _txt(registro.get("Codigo"))
        digitos = re.sub(r"\D", "", codigo)
        if not digitos:
            continue
        try:
            fim = datetime.strptime(_txt(registro.get("Data_Fim")), "%d/%m/%Y").date()
        except ValueError:
            fim = None
        itens.append(
            {
                "ncm": digitos,
                "codigo": codigo,
                # a fonte traz marcacao HTML no meio da descricao (ex.: <i>UHT</i>)
                "descricao": _txt(
                    html.unescape(re.sub(r"<[^>]+>", "", str(registro.get("Descricao") or "")))
                ).lstrip("- ").strip(),
                "nivel": len(digitos),
                "vigente": fim is None or fim >= hoje,
                "data_inicio": _txt(registro.get("Data_Inicio")),
                "data_fim": _txt(registro.get("Data_Fim")),
                "ato": " ".join(
                    filter(
                        None,
                        [
                            _txt(registro.get("Tipo_Ato_Ini")),
                            _txt(registro.get("Numero_Ato_Ini")),
                            _txt(registro.get("Ano_Ato_Ini")),
                        ],
                    )
                ),
            }
        )
    return itens


# ------------------------------------------------- Anexos da LC 214/2025 (NCM)

RE_REF_NCM = re.compile(r"(?<![\d.,])(\d{2}\.\d{2}(?:\.\d{1,2}){0,2}|\d{4}\.\d{1,2}(?:\.\d{1,2})?)(?![\d])")
RE_POSICAO_SECA = re.compile(r"(?<![\d.,/])(\d{4})(?![\d/]|\.\d)")
TERMOS_EXCECAO = ("exceto", "excluid", "excluíd", "ressalvad", "salvo")


class _LeitorAnexos(HTMLParser):
    """Coleta paragrafos e linhas de tabela do texto da LC 214/2025, em ordem."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.eventos: list[tuple[str, object]] = []
        self._pilha_tabela = 0
        self._celulas: list[str] | None = None
        self._buffer: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._pilha_tabela += 1
        elif tag == "tr" and self._pilha_tabela:
            self._celulas = []
        elif tag in ("td", "th") and self._celulas is not None:
            self._descarrega_celula()
        elif tag == "p" and self._celulas is None:
            self._descarrega_paragrafo()
        elif tag == "br":
            self._buffer.append(" ")

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._celulas is not None:
            self._descarrega_celula()
        elif tag == "tr" and self._celulas is not None:
            if any(c for c in self._celulas):
                self.eventos.append(("linha", self._celulas))
            self._celulas = None
        elif tag == "table":
            self._pilha_tabela = max(0, self._pilha_tabela - 1)
        elif tag == "p" and self._celulas is None:
            self._descarrega_paragrafo()

    def handle_data(self, data):
        self._buffer.append(data)

    def _texto(self) -> str:
        texto = _txt("".join(self._buffer))
        self._buffer = []
        return texto

    def _descarrega_celula(self) -> None:
        texto = self._texto()
        if self._celulas is not None and texto:
            self._celulas.append(texto)

    def _descarrega_paragrafo(self) -> None:
        texto = self._texto()
        if texto:
            self.eventos.append(("paragrafo", texto))


def _coluna_de_codigos(celula: str) -> bool:
    """Diz se a celula e a coluna 'NCM/SH' (so codigos, sem texto corrido)."""
    if not celula:
        return False
    resto = RE_REF_NCM.sub(" ", celula)
    resto = re.sub(r"[\s;,.eE/()\-]|\bex\b", "", resto)
    return bool(RE_REF_NCM.search(celula)) and len(resto) <= 3


def _referencias_ncm(texto: str) -> tuple[list[str], list[str]]:
    """Separa as referencias de NCM citadas como incluidas e como excecao.

    Alem dos codigos pontuados (1006.40.00, 0306.1, 87.03), aceita posicoes
    "secas" de 4 digitos (ex.: "8802, exceto o codigo 8802.60.00") quando o
    trecho ja contem alguma referencia pontuada - assim numeros soltos de outros
    contextos nao viram NCM.
    """
    incluidas: list[str] = []
    excluidas: list[str] = []

    def classifica(inicio: int, digitos: str) -> None:
        anterior = texto[max(0, inicio - 260):inicio].lower()
        corte = max(anterior.rfind(";"), anterior.rfind(":"))
        trecho = anterior[corte + 1:]
        alvo = excluidas if any(t in trecho for t in TERMOS_EXCECAO) else incluidas
        if digitos not in alvo:
            alvo.append(digitos)

    achou_pontuado = False
    for m in RE_REF_NCM.finditer(texto):
        digitos = m.group(1).replace(".", "")
        if len(digitos) < 4:
            continue
        achou_pontuado = True
        classifica(m.start(), digitos)

    # listas do tipo "2401; 2402; 2403" tambem sao posicoes da NCM
    lista_de_posicoes = re.search(r"\d{4}\s*;\s*\d{4}", texto) is not None
    if achou_pontuado or lista_de_posicoes:
        for m in RE_POSICAO_SECA.finditer(texto):
            classifica(m.start(), m.group(1))

    return incluidas, excluidas


def build_lc214_anexos() -> list[dict]:
    """Extrai os itens dos anexos da LC 214/2025 e as referencias de NCM."""
    arq = _unico("legislacao/LC_214_2025_camara.html")
    leitor = _LeitorAnexos()
    leitor.feed(arq.read_text(encoding="utf-8", errors="replace"))

    itens: list[dict] = []
    anexo = titulo = rotulo = ""
    aguardando_titulo = False
    for tipo, conteudo in leitor.eventos:
        if tipo == "paragrafo":
            texto = str(conteudo)
            m = re.fullmatch(r"ANEXO\s+([IVXL]+)\s*(?:\(.*\))?", texto)
            if m:
                anexo, titulo, aguardando_titulo = m.group(1), "", True
            elif aguardando_titulo and len(texto) > 20:
                titulo = texto
                aguardando_titulo = False
            continue

        celulas = list(conteudo)  # type: ignore[arg-type]
        if not anexo or not celulas:
            continue
        if len(celulas) == 1:
            # anexos de coluna unica (ex.: XVII): a linha alterna rotulo e lista
            unica = celulas[0]
            tem_codigo = RE_REF_NCM.search(unica) or re.search(r"\d{4}\s*;\s*\d{4}", unica)
            if not tem_codigo:
                rotulo = unica[:60]
                continue
            celulas = [rotulo, unica]
        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})*", celulas[0]):
            ordem, corpo = celulas[0], celulas[1:]
        else:
            # anexos cuja primeira coluna e um rotulo (ex.: Anexo XVII)
            ordem, corpo = celulas[0][:60], celulas

        descricao = " ".join(corpo)
        if len(corpo) >= 2 and _coluna_de_codigos(corpo[-1]):
            # layout "ITEM | DESCRICAO | NCM/SH": a coluna de codigos e a inclusao;
            # a descricao so contribui com as excecoes que ela enumera.
            incluidas, _ = _referencias_ncm(corpo[-1])
            _, excluidas = _referencias_ncm(" ".join(corpo[:-1]))
            excluidas = [d for d in excluidas if d not in incluidas]
        else:
            incluidas, excluidas = _referencias_ncm(descricao)
        if not incluidas and not excluidas:
            continue
        itens.append(
            {
                "anexo": anexo,
                "anexo_titulo": titulo,
                "item": ordem,
                "descricao": descricao,
                "ncm_incluidos": incluidas,
                "ncm_excecoes": excluidas,
            }
        )
    return itens


# ---------------------------------------------------------------------- CNAE

SECOES_BENS = {"A", "B", "C", "G"}
SECOES_SERVICO = set("DEFHIJKLMNOPQRSU")


def build_cnae() -> list[dict]:
    arq = _unico("cnae/ibge_cnae_subclasses.json")
    bruto = json.loads(arq.read_text(encoding="utf-8"))
    itens = []
    for sub in bruto:
        classe = sub.get("classe") or {}
        grupo = classe.get("grupo") or {}
        divisao = grupo.get("divisao") or {}
        secao = divisao.get("secao") or {}
        codigo = re.sub(r"\D", "", str(sub["id"]))
        itens.append(
            {
                "cnae": codigo,
                "cnae_formatado": f"{codigo[0:4]}-{codigo[4]}/{codigo[5:7]}",
                "descricao": _txt(sub.get("descricao")),
                "classe": _txt(classe.get("id")),
                "classe_descricao": _txt(classe.get("descricao")),
                "divisao": _txt(divisao.get("id")),
                "divisao_descricao": _txt(divisao.get("descricao")),
                "secao": _txt(secao.get("id")),
                "secao_descricao": _txt(secao.get("descricao")),
                "observacoes": [_txt(o) for o in (sub.get("observacoes") or [])],
            }
        )
    return itens


# --------------------------------------------------- CNAE x item da LC 116/03


def _itens_de_texto(texto: str) -> list[str]:
    """Extrai codigos de item LC116 ('1.01', '01.01', '0101') normalizados."""
    achados = []
    for m in re.finditer(r"\b(\d{1,2})[.\-]?(\d{2})\b", texto):
        item = f"{int(m.group(1)):02d}.{m.group(2)}"
        achados.append(item)
    return achados


def parse_anapolis(pdf) -> list[tuple[str, str]]:
    """Portaria 463/2025 de Anapolis-GO: colunas CNAE mascara | CNAE | desc | item | ..."""
    pares = []
    for pagina in pdf.pages:
        for tabela in pagina.extract_tables() or []:
            for linha in tabela:
                celulas = [_txt(c) for c in linha]
                if len(celulas) < 5:
                    continue
                cnae = re.sub(r"\D", "", celulas[1])
                if len(cnae) != 7:
                    continue
                item = _itens_de_texto(celulas[3])
                if not item:
                    item = _itens_de_texto(celulas[4])
                if item:
                    pares.append((cnae, item[0]))
    return pares


def parse_salvador(pdf) -> list[tuple[str, str]]:
    """Tabela CNAE x Lista de Servicos de Salvador-BA (texto corrido)."""
    pares = []
    padrao = re.compile(r"(\d{4}-\d/\d{2})\s+(\d{2}\.\d{2})")
    for pagina in pdf.pages:
        for m in padrao.finditer(pagina.extract_text() or ""):
            pares.append((re.sub(r"\D", "", m.group(1)), m.group(2)))
    return pares


def build_cnae_lc116(itens_validos: set[str]) -> list[dict]:
    import pdfplumber

    fontes = [
        ("Anapolis-GO (Portaria 463/2025)", "cnae-lc116/Anapolis_*.pdf", parse_anapolis),
        ("Salvador-BA (SEFAZ, tabela CNAE x Lista)", "cnae-lc116/Salvador_*.pdf", parse_salvador),
    ]
    mapa: dict[tuple[str, str], set[str]] = {}
    for rotulo, glob, parser in fontes:
        try:
            arq = _unico(glob)
        except SystemExit:
            print(f"  ! fonte ausente: {glob}")
            continue
        with pdfplumber.open(arq) as pdf:
            pares = parser(pdf)
        validos = 0
        for cnae, item in pares:
            if item not in itens_validos:
                continue
            mapa.setdefault((cnae, item), set()).add(rotulo)
            validos += 1
        print(f"  {rotulo}: {validos} vinculos aproveitados de {len(pares)} lidos")

    saida = []
    for (cnae, item), rotulos in sorted(mapa.items()):
        saida.append({"cnae": cnae, "item": item, "fontes": sorted(rotulos)})
    return saida


# ---------------------------------------------------------------------- main


def gravar(nome: str, dados) -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    caminho = DATA / nome
    caminho.write_text(
        json.dumps(dados, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    tamanho = len(dados) if isinstance(dados, list) else "-"
    print(f"  {nome}: {tamanho} registros")


def main() -> int:
    print("Gerando datasets a partir de KB/ ...")

    lc116 = build_lc116()
    gravar("lc116.json", lc116)

    gravar("nbs.json", build_nbs())
    print("Lendo as notas explicativas da NBS (PDF) ...")
    gravar("nebs.json", build_nebs())
    gravar("indop.json", build_indop())

    classes, cst = build_cclasstrib()
    gravar("cclasstrib.json", classes)
    gravar("cst.json", cst)

    gravar("anexo_viii.json", build_anexo_viii())
    gravar("cnae.json", build_cnae())
    gravar("ncm.json", build_ncm())
    gravar("lc214_anexos.json", build_lc214_anexos())

    print("Lendo tabelas municipais CNAE x LC 116 ...")
    gravar("cnae_lc116.json", build_cnae_lc116({i["item"] for i in lc116}))

    catalogo = {}
    catalogo_arq = KB / "fontes.json"
    if catalogo_arq.exists():
        for f in json.loads(catalogo_arq.read_text(encoding="utf-8")):
            catalogo[f["arquivo"]] = f

    fontes = []
    for p in sorted(KB.rglob("*")):
        if not p.is_file() or p.suffix.lower() not in {".xlsx", ".csv", ".pdf", ".json", ".html"}:
            continue
        if p.name == "fontes.json":
            continue
        rel = str(p.relative_to(RAIZ)).replace("\\", "/")
        registro = dict(catalogo.get(rel, {"arquivo": rel}))
        try:
            registro["sha256"] = _sha256(p)
        except PermissionError:
            registro["sha256"] = "indisponivel"
        registro["bytes"] = p.stat().st_size
        registro["baixado_em"] = datetime.fromtimestamp(p.stat().st_mtime).isoformat(
            timespec="seconds"
        )
        registro["catalogado"] = rel in catalogo
        fontes.append(registro)

    manifesto = {
        "gerado_em": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "fontes": fontes,
    }
    gravar("manifesto.json", manifesto)
    print("OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
