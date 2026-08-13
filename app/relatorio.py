"""Geracao dos arquivos de saida (XLSX, CSV e JSON) da correlacao."""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .correlacao import linhas_detalhe, resumo
from .dados import base

CABECALHO = [
    "CNAE", "Tipo", "Descrição CNAE", "Seção CNAE", "Natureza",
    "Item LC 116/03", "Descrição item", "NBS", "Descrição NBS",
    "indOP", "Descrição indOP", "Local de incidência",
    "cClassTrib", "Nome cClassTrib", "CST IBS/CBS",
    "Confiança", "Origem", "Alertas",
]

_AZUL = PatternFill("solid", fgColor="1F3864")
_CINZA = PatternFill("solid", fgColor="F2F2F2")


def _ajusta(ws, larguras: dict[int, int]) -> None:
    for col, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


def gerar_xlsx(cartao: dict, resultados: list[dict]) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Correlacao"
    ws.append(CABECALHO)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _AZUL
        celula.alignment = Alignment(vertical="center", wrap_text=True)
    for linha in linhas_detalhe(resultados):
        ws.append([linha.get(c, "") for c in CABECALHO])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _ajusta(ws, {1: 12, 2: 11, 3: 45, 4: 30, 5: 20, 6: 13, 7: 45, 8: 15, 9: 45,
                 10: 10, 11: 40, 12: 32, 13: 12, 14: 40, 15: 12, 16: 12, 17: 45, 18: 60})

    wsr = wb.create_sheet("Resumo")
    r = resumo(resultados)
    dados = [
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("CNPJ", cartao.get("cnpj", "")),
        ("Nome empresarial", cartao.get("nome_empresarial", "")),
        ("Nome fantasia", cartao.get("nome_fantasia", "")),
        ("Municipio/UF", f"{cartao.get('municipio','')} / {cartao.get('uf','')}"),
        ("Natureza juridica", cartao.get("natureza_juridica", "")),
        ("Situacao cadastral", cartao.get("situacao", "")),
        ("", ""),
        ("CNAEs analisados", r["cnaes"]),
        ("CNAEs de servico (vinculo oficial)", r["servicos"]),
        ("CNAEs de servico (sugestao a validar)", r["servicos_sugeridos"]),
        ("CNAEs de operacao com bens", r["bens"]),
        ("Itens da LC 116/03 distintos", r["itens_lc116"]),
        ("Codigos NBS distintos", r["nbs"]),
        ("Codigos cClassTrib distintos", r["cclasstrib"]),
    ]
    for chave, valor in dados:
        wsr.append([chave, valor])
    for linha in wsr.iter_rows(min_col=1, max_col=1):
        linha[0].font = Font(bold=True)
    _ajusta(wsr, {1: 40, 2: 60})

    wsf = wb.create_sheet("Fontes")
    wsf.append(["Arquivo (KB)", "Titulo", "Orgao", "Versao", "URL", "SHA-256", "Baixado em"])
    for celula in wsf[1]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for fonte in base().versao_fontes():
        wsf.append([
            fonte.get("arquivo", ""), fonte.get("titulo", ""), fonte.get("orgao", ""),
            fonte.get("versao", ""), fonte.get("url", ""), fonte.get("sha256", "")[:16],
            fonte.get("baixado_em", ""),
        ])
    _ajusta(wsf, {1: 55, 2: 55, 3: 35, 4: 28, 5: 70, 6: 20, 7: 20})

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


CABECALHO_NCM = [
    "NCM", "Descrição NCM", "Vigente", "Regime", "Anexo LC 214/2025",
    "Item do anexo", "Texto do anexo", "cClassTrib", "Nome cClassTrib",
    "CST IBS/CBS", "Descrição CST", "Redução IBS (%)", "Redução CBS (%)",
    "Base legal", "indOP aplicáveis", "Observação",
]


def gerar_xlsx_ncm(consulta: dict) -> bytes:
    from .ncm import linhas_detalhe as linhas_ncm

    wb = Workbook()
    ws = wb.active
    ws.title = "NCM"
    ws.append(CABECALHO_NCM)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _AZUL
        celula.alignment = Alignment(vertical="center", wrap_text=True)
    for linha in linhas_ncm(consulta):
        ws.append([linha.get(c, "") for c in CABECALHO_NCM])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _ajusta(ws, {1: 14, 2: 50, 3: 9, 4: 26, 5: 18, 6: 14, 7: 70, 8: 12, 9: 45,
                 10: 12, 11: 28, 12: 14, 13: 14, 14: 20, 15: 26, 16: 45})

    wsi = wb.create_sheet("indOP")
    wsi.append(["indOP", "Tipo de operação", "Característica", "Local do fornecimento",
                "Dispositivo LC 214/2025"])
    for celula in wsi[1]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for indop in consulta["indops"]:
        wsi.append([indop["indop"], indop["tipo_operacao"], indop["caracteristica"],
                    indop["local_fornecimento"], indop["dispositivo_lc214"]])
    _ajusta(wsi, {1: 10, 2: 30, 3: 55, 4: 55, 5: 25})

    wsh = wb.create_sheet("Hierarquia e alertas")
    wsh.append(["Nível", "Código", "Descrição"])
    for celula in wsh[1]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for nivel in consulta["hierarquia"]:
        wsh.append([nivel["nivel"], nivel["codigo"], nivel["descricao"]])
    wsh.append([])
    wsh.append(["Alertas"])
    for alerta in consulta["alertas"]:
        wsh.append(["", alerta])
    _ajusta(wsh, {1: 8, 2: 16, 3: 90})

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_csv_ncm(consulta: dict) -> bytes:
    from .ncm import linhas_detalhe as linhas_ncm

    buffer = io.StringIO()
    escritor = csv.DictWriter(buffer, fieldnames=CABECALHO_NCM, delimiter=";",
                              extrasaction="ignore")
    escritor.writeheader()
    for linha in linhas_ncm(consulta):
        escritor.writerow(linha)
    return buffer.getvalue().encode("utf-8-sig")


def gerar_json_ncm(consulta: dict) -> bytes:
    payload = {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "consulta": consulta,
        "fontes": base().versao_fontes(),
    }
    return json.dumps(payload, ensure_ascii=False, indent=1).encode("utf-8")


def gerar_csv(resultados: list[dict]) -> bytes:
    buffer = io.StringIO()
    escritor = csv.DictWriter(buffer, fieldnames=CABECALHO, delimiter=";", extrasaction="ignore")
    escritor.writeheader()
    for linha in linhas_detalhe(resultados):
        escritor.writerow(linha)
    return buffer.getvalue().encode("utf-8-sig")


def gerar_json(cartao: dict, resultados: list[dict]) -> bytes:
    payload = {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "empresa": cartao,
        "resumo": resumo(resultados),
        "correlacoes": resultados,
        "fontes": base().versao_fontes(),
    }
    return json.dumps(payload, ensure_ascii=False, indent=1).encode("utf-8")


# ------------------------------------------------------------------ cesta

def _aba_fontes(wb) -> None:
    ws = wb.create_sheet("Fontes")
    ws.append(["Arquivo (KB)", "Título", "Órgão", "Versão", "URL", "SHA-256", "Baixado em"])
    for celula in ws[1]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for fonte in base().versao_fontes():
        ws.append([
            fonte.get("arquivo", ""), fonte.get("titulo", ""), fonte.get("orgao", ""),
            fonte.get("versao", ""), fonte.get("url", ""), fonte.get("sha256", "")[:16],
            fonte.get("baixado_em", ""),
        ])
    _ajusta(ws, {1: 55, 2: 55, 3: 35, 4: 28, 5: 70, 6: 20, 7: 20})


def gerar_xlsx_cesta(cesta: dict) -> bytes:
    """Planilha consolidada da cesta: linhas, resumo, consultas e fontes."""
    from .cesta import COLUNAS, linhas as linhas_cesta, resumo as resumo_cesta

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    ws.append(COLUNAS)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = _AZUL
        celula.alignment = Alignment(vertical="center", wrap_text=True)
    for linha in linhas_cesta(cesta):
        ws.append([linha.get(c, "") for c in COLUNAS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _ajusta(ws, {1: 26, 2: 16, 3: 14, 4: 45, 5: 14, 6: 15, 7: 45, 8: 22, 9: 12,
                 10: 42, 11: 12, 12: 13, 13: 13, 14: 22, 15: 60, 16: 22, 17: 60})

    r = resumo_cesta(cesta)
    wsr = wb.create_sheet("Resumo")
    for chave, valor in [
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Cesta criada em", cesta.get("criada_em", "")),
        ("", ""),
        ("Consultas", r["consultas"]),
        ("  cartão CNPJ / CNAE", r["consultas_cnae"]),
        ("  NCM", r["consultas_ncm"]),
        ("Linhas consolidadas", r["linhas"]),
        ("Códigos distintos", r["codigos"]),
        ("cClassTrib distintos", r["cclasstrib"]),
        ("", ""),
        ("Linhas com alíquota zero (redução 100%)", r["aliquota_zero"]),
        ("Linhas com redução de 60%", r["reducao_60"]),
        ("Linhas com redução de 30%", r["reducao_30"]),
        ("Linhas de tributação integral (CST 000)", r["integral"]),
        ("Linhas a validar (similaridade)", r["a_validar"]),
        ("Linhas com Imposto Seletivo", r["imposto_seletivo"]),
        ("Linhas em cláusula de exceção", r["excecoes"]),
    ]:
        wsr.append([chave, valor])
    wsr.append([])
    wsr.append(["CST", "Descrição", "Linhas"])
    for celula in wsr[wsr.max_row]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for grupo in r["top_cst"]:
        wsr.append([grupo["cst"], grupo["descricao"], grupo["linhas"]])
    wsr.append([])
    wsr.append(["cClassTrib", "Nome", "Linhas"])
    for celula in wsr[wsr.max_row]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for grupo in r["top_cclasstrib"]:
        wsr.append([grupo["cclasstrib"], grupo["nome"], grupo["linhas"]])
    for linha in wsr.iter_rows(min_col=1, max_col=1):
        if linha[0].value and str(linha[0].value)[0].isupper():
            linha[0].font = Font(bold=True)
    _ajusta(wsr, {1: 42, 2: 60, 3: 12})

    wsc = wb.create_sheet("Consultas")
    wsc.append(["#", "Tipo", "Consulta", "Descrição", "Detalhe", "Adicionada em"])
    for celula in wsc[1]:
        celula.font = Font(bold=True)
        celula.fill = _CINZA
    for ordem, item in enumerate(cesta["itens"], start=1):
        wsc.append([
            ordem,
            "Cartão CNPJ" if item["tipo"] == "cnae" else "NCM",
            item["rotulo"], item["titulo"], item["detalhe"], item["adicionado_em"],
        ])
    _ajusta(wsc, {1: 5, 2: 14, 3: 26, 4: 55, 5: 30, 6: 18})

    _aba_fontes(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_csv_cesta(cesta: dict) -> bytes:
    from .cesta import COLUNAS, linhas as linhas_cesta

    buffer = io.StringIO()
    escritor = csv.DictWriter(buffer, fieldnames=COLUNAS, delimiter=";",
                              extrasaction="ignore")
    escritor.writeheader()
    for linha in linhas_cesta(cesta):
        escritor.writerow(linha)
    return buffer.getvalue().encode("utf-8-sig")


def gerar_json_cesta(cesta: dict) -> bytes:
    from .cesta import linhas as linhas_cesta, resumo as resumo_cesta

    payload = {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "cesta": {
            "criada_em": cesta.get("criada_em", ""),
            "consultas": [
                {k: v for k, v in item.items() if k != "linhas"}
                for item in cesta["itens"]
            ],
        },
        "resumo": resumo_cesta(cesta),
        "linhas": linhas_cesta(cesta),
        "fontes": base().versao_fontes(),
    }
    return json.dumps(payload, ensure_ascii=False, indent=1).encode("utf-8")
